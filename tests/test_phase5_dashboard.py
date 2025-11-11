"""Tests for Phase 5 dashboard generation."""
from __future__ import annotations

import sys
from pathlib import Path
from textwrap import dedent

import pandas as pd
from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))

from src.phase5_dashboard import run_phase5


def _write_dataframe(df: pd.DataFrame, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    df.to_csv(path, index=False)


def _build_config(tmp_path: Path) -> Path:
    data_dir = tmp_path / "data"
    processed_dir = data_dir / "processed"
    scorecard_dir = processed_dir / "scorecards"
    commentary_dir = processed_dir / "commentary"
    forecast_dir = processed_dir / "forecasts"
    dashboard_dir = processed_dir / "dashboard"
    logs_dir = tmp_path / "logs"
    logs_dir.mkdir(parents=True, exist_ok=True)

    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        dedent(
            f"""
            paths:
              raw_dir: {data_dir / 'raw'}
              processed_dir: {processed_dir}
              logs_dir: {logs_dir}
              calendar_map: {data_dir / 'reference' / 'calendar_map.csv'}
              normalized_output: {processed_dir / 'normalized_phase1.csv'}
              scorecard_vendor_csv: {scorecard_dir / 'vendor.csv'}
              scorecard_asin_csv: {scorecard_dir / 'asin.csv'}
              forecasts_dir: {forecast_dir}
              forecast_vendor_csv: {forecast_dir / 'vendor.csv'}
              forecast_asin_csv: {forecast_dir / 'asin.csv'}
              dashboard_dir: {dashboard_dir}
            dashboard:
              output_dir: {dashboard_dir}
              file_prefix: TestDashboard
              sheet_names:
                summary: Summary
                vendor_scorecard: Vendor Scorecards
                asin_scorecard: ASIN Scorecards
                commentary: Commentary
                forecasts: Forecasts
              sheet_order:
                - Summary
                - Vendor Scorecards
                - ASIN Scorecards
                - Commentary
                - Forecasts
              formatting:
                header_fill: "1F4E78"
                header_font_color: "FFFFFF"
                body_font: Calibri
                header_font_size: 12
                body_font_size: 11
                precision: 2
                body_alignment:
                  horizontal: left
                  vertical: center
                  wrap_text: true
                row_height: 20
                header_row_height: 24
                borders:
                  style: thin
                  color: D9D9D9
                column_widths:
                  default: 18
                  overrides:
                    Commentary_Insight: 50
                    Forecast_Potential_Gain: 20
                score_color_scale:
                  start_color: "F8696B"
                  mid_color: "FFEB84"
                  end_color: "63BE7B"
                  mid_value: 700
                  end_value: 950
                icon_sets:
                  - column: Composite_Score
                    style: 3TrafficLights1
                    type: num
                    values:
                      - 0
                      - 700
                      - 900
                  - column: Forecast_Potential_Gain
                    style: 3Arrows
                    type: num
                    values:
                      - 0
                      - 5
                      - 10
                commentary_fills:
                  performance_summary: "C6EFCE"
                  opportunity: "FFEB9C"
                  risk: "FFC7CE"
                  default: "FFFFFF"
                summary_number_format: "#,##0.00"
            """
        ).strip()
        + "\n",
        encoding="utf-8",
    )
    return config_path


def _create_inputs(tmp_path: Path) -> None:
    processed_dir = tmp_path / "data" / "processed"
    processed_dir.mkdir(parents=True, exist_ok=True)

    normalized = pd.DataFrame(
        [
            {
                "vendor_code": "V1",
                "asin": "A1",
                "metric": "GMS($)",
                "week_label": "2025W36",
                "week_start": "2025-09-01",
                "value": 100.0,
            },
            {
                "vendor_code": "V2",
                "asin": "A2",
                "metric": "GMS($)",
                "week_label": "2025W36",
                "week_start": "2025-09-01",
                "value": 80.0,
            },
            {
                "vendor_code": "V1",
                "asin": "A1",
                "metric": "GMS($)",
                "week_label": "2025W37",
                "week_start": "2025-09-08",
                "value": 110.0,
            },
        ]
    )
    _write_dataframe(normalized, processed_dir / "normalized_phase1.csv")

    vendor_scorecard = pd.DataFrame(
        [
            {
                "entity_type": "vendor",
                "vendor_code": "V1",
                "vendor_name": "Vendor One",
                "metric": "GMS($)",
                "Composite_Score": 920.0,
                "Improvement_Flag": "Watch",
                "Raw_Value": 110.0,
            },
            {
                "entity_type": "vendor",
                "vendor_code": "V2",
                "vendor_name": "Vendor Two",
                "metric": "GMS($)",
                "Composite_Score": 640.0,
                "Improvement_Flag": "Critical",
                "Raw_Value": 80.0,
            },
        ]
    )
    _write_dataframe(vendor_scorecard, processed_dir / "scorecards" / "vendor.csv")

    asin_scorecard = pd.DataFrame(
        [
            {
                "entity_type": "asin",
                "vendor_code": "V1",
                "vendor_name": "Vendor One",
                "asin": "A1",
                "metric": "GMS($)",
                "Composite_Score": 910.0,
                "Raw_Value": 110.0,
            },
            {
                "entity_type": "asin",
                "vendor_code": "V2",
                "vendor_name": "Vendor Two",
                "asin": "A2",
                "metric": "GMS($)",
                "Composite_Score": 600.0,
                "Raw_Value": 80.0,
            },
        ]
    )
    _write_dataframe(asin_scorecard, processed_dir / "scorecards" / "asin.csv")

    commentary = pd.DataFrame(
        [
            {
                "entity_type": "vendor",
                "vendor_code": "V1",
                "asin": None,
                "week_label": "2025W37",
                "metric": "GMS($)",
                "Commentary_Type": "performance_summary",
                "Commentary_Text": "V1 grew GMS by 10% WoW.",
                "Change_Class": "strong_increase",
                "Change_Pct": 0.1,
                "Current_Value": 110.0,
                "Previous_Value": 100.0,
            },
            {
                "entity_type": "asin",
                "vendor_code": "V1",
                "asin": "A1",
                "week_label": "2025W37",
                "metric": "GMS($)",
                "Commentary_Type": "opportunity",
                "Commentary_Text": "A1 has room to lift ASP.",
                "Change_Class": "flat",
                "Change_Pct": 0.0,
                "Current_Value": 110.0,
                "Previous_Value": 110.0,
            },
        ]
    )
    _write_dataframe(commentary, processed_dir / "commentary" / "commentary.csv")

    vendor_forecast = pd.DataFrame(
        [
            {
                "entity_type": "vendor",
                "vendor_code": "V1",
                "asin": None,
                "metric": "GMS($)",
                "forecast_week": "2025-09-15",
                "forecast_value": 115.0,
                "potential_gain": 12.0,
            },
            {
                "entity_type": "vendor",
                "vendor_code": "V2",
                "asin": None,
                "metric": "GMS($)",
                "forecast_week": "2025-09-15",
                "forecast_value": 85.0,
                "potential_gain": 6.0,
            },
        ]
    )
    _write_dataframe(vendor_forecast, processed_dir / "forecasts" / "vendor.csv")

    asin_forecast = pd.DataFrame(
        [
            {
                "entity_type": "asin",
                "vendor_code": "V1",
                "asin": "A1",
                "metric": "GMS($)",
                "forecast_week": "2025-09-15",
                "forecast_value": 116.0,
                "potential_gain": 11.0,
            },
        ]
    )
    _write_dataframe(asin_forecast, processed_dir / "forecasts" / "asin.csv")


def test_dashboard_generation_creates_expected_sheets(tmp_path: Path) -> None:
    config_path = _build_config(tmp_path)
    _create_inputs(tmp_path)

    workbook_path, summary = run_phase5(config_path=str(config_path))

    assert workbook_path.exists()
    wb = load_workbook(workbook_path)
    expected_order = ["Summary", "Vendor Scorecards", "ASIN Scorecards", "Commentary", "Forecasts"]
    assert wb.sheetnames[: len(expected_order)] == expected_order

    vendor_sheet = wb["Vendor Scorecards"]
    headers = [cell.value for cell in vendor_sheet[1]]
    assert "Commentary_Insight" in headers
    assert "Forecast_Potential_Gain" in headers
    comment_col = headers.index("Commentary_Insight") + 1
    forecast_col = headers.index("Forecast_Potential_Gain") + 1
    assert vendor_sheet.cell(row=2, column=comment_col).value == "V1 grew GMS by 10% WoW."
    assert vendor_sheet.cell(row=2, column=forecast_col).value == 12.0
    assert len(vendor_sheet.conditional_formatting) > 0
    header_font = vendor_sheet.cell(row=1, column=1).font
    assert header_font.bold
    assert header_font.size == 12
    body_cell = vendor_sheet.cell(row=2, column=1)
    assert body_cell.font.size == 11
    assert body_cell.alignment.horizontal == "left"
    assert vendor_sheet.row_dimensions[2].height == 20
    assert body_cell.border.left.style == "thin"
    icon_rules = []
    cf_rules = getattr(vendor_sheet.conditional_formatting, "_cf_rules", {})
    for rules in cf_rules.values():
        icon_rules.extend([rule for rule in rules if getattr(rule, "type", None) == "iconSet"])
    assert icon_rules

    summary_sheet = wb["Summary"]
    values = {summary_sheet.cell(row=row, column=1).value: summary_sheet.cell(row=row, column=2).value for row in range(3, 9)}
    assert values["Total Vendors"] == 2
    assert round(values["Average Vendor Score"], 1) == 780.0
    assert values["Data Freshness"] == "2025-09-08"

    assert summary["Total Vendors"] == 2
    assert summary["Highest Forecast Gain"] == 12.0

