"""Phase 4 Output Layer: machine-readable and Excel deliverables.

This module provides:
- build_phase4_run_dir: create timestamped directories under outputs/phase4/
- collect_git_commit_hash: try to read current git short hash
- write_metadata_json: persist metadata.json in run folder
- write_machine_outputs: write Parquet artifacts for long/wide metrics and scores
- write_excel_outputs: produce a diagnostics workbook (openpyxl)

All functions are side-effect limited to path arguments and raise exceptions on
hard failures (callers decide error mode). Parquet writes use pyarrow.
"""
from __future__ import annotations

from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Dict, Optional, List, Tuple

import json
import os
import shutil
import subprocess

import pandas as pd
import numpy as np
import hashlib


def build_phase4_run_dir(base_dir: Path | str = "outputs/phase4") -> Dict[str, Path]:
    base = Path(base_dir)
    base.mkdir(parents=True, exist_ok=True)
    ts = datetime.now().strftime("run_%Y%m%d_%H%M%S")
    run_dir = base / ts
    excel_dir = run_dir / "excel"
    machine_dir = run_dir / "machine"
    excel_dir.mkdir(parents=True, exist_ok=True)
    machine_dir.mkdir(parents=True, exist_ok=True)
    return {"run": run_dir, "excel": excel_dir, "machine": machine_dir}


def collect_git_commit_hash(cwd: Optional[Path] = None) -> Optional[str]:
    try:
        result = subprocess.run(
            ["git", "rev-parse", "--short", "HEAD"],
            cwd=str(cwd) if cwd else None,
            capture_output=True,
            text=True,
            check=True,
        )
        h = result.stdout.strip()
        return h or None
    except Exception:
        return None


def write_metadata_json(path: Path, payload: Dict[str, object]) -> Path:
    path = Path(path)
    with path.open("w", encoding="utf-8") as fh:
        json.dump(payload, fh, indent=2, ensure_ascii=False)
    return path


def _ensure_parquet(df: pd.DataFrame, out_path: Path) -> Path:
    out_path = Path(out_path)
    df.to_parquet(out_path, index=False)
    return out_path


def _schema_signature(df: pd.DataFrame) -> str:
    """Return a stable string representation of the DataFrame schema (name:dtype)."""
    cols = [(c, str(dt)) for c, dt in zip(df.columns, df.dtypes)]
    return "|".join(f"{c}:{t}" for c, t in cols)


def schema_checksum(df: pd.DataFrame) -> str:
    """Compute md5 checksum of schema signature for metadata purposes."""
    sig = _schema_signature(df).encode("utf-8")
    return "md5:" + hashlib.md5(sig).hexdigest()


def _is_numeric(series: pd.Series) -> bool:
    return pd.api.types.is_numeric_dtype(series)


def _required_long_columns() -> List[str]:
    return [
        "entity_id",
        "vendor_code",
        "Week_Order",
        "metric",
        "value",
        "raw_value",
        "std_value",
        "volatility",
        "trend_slope",
        "trend_strength",
        "completeness",
        "recency_weeks",
        "variance",
        "trend_quality",
        "weight_effective",
        "weight_multiplier",
        "base_weight",
    ]


def _extract_metrics_from_wide(df_wide: pd.DataFrame) -> List[str]:
    return sorted({col.split("std_")[1] for col in df_wide.columns if col.startswith("std_")})


def validate_output_schema(df_long: pd.DataFrame, df_wide: pd.DataFrame) -> None:
    """Validate machine output schemas for Patch-3 compliance.

    Raises ValueError with descriptive messages when checks fail.
    """
    # Long checks
    missing_long = [c for c in _required_long_columns() if c not in df_long.columns]
    if missing_long:
        raise ValueError(f"metrics_long missing required columns: {missing_long}")
    # Identity not null
    for col in ["entity_id", "vendor_code", "Week_Order", "metric"]:
        if df_long[col].isna().any():
            raise ValueError(f"metrics_long has nulls in identity column '{col}'")
    # Dtype checks
    numeric_long_cols = [
        "value",
        "raw_value",
        "std_value",
        "volatility",
        "trend_slope",
        "trend_strength",
        "completeness",
        "recency_weeks",
        "variance",
        "trend_quality",
        "weight_effective",
        "weight_multiplier",
        "base_weight",
    ]
    for col in numeric_long_cols:
        if not _is_numeric(df_long[col]):
            raise ValueError(f"metrics_long column '{col}' must be numeric, got {df_long[col].dtype}")

    # Wide checks
    if "score_composite" not in df_wide.columns:
        raise ValueError("metrics_wide missing 'score_composite'")
    metrics = _extract_metrics_from_wide(df_wide)
    if not metrics:
        raise ValueError("metrics_wide has no 'std_<metric>' columns")
    for m in metrics:
        for prefix in ("std_", "contrib_", "weight_", "weightmul_", "base_weight_"):
            col = f"{prefix}{m}"
            if col not in df_wide.columns:
                raise ValueError(f"metrics_wide missing audit column '{col}'")
            if prefix in ("std_", "contrib_", "weight_", "weightmul_", "base_weight_") and not _is_numeric(df_wide[col]):
                raise ValueError(f"metrics_wide column '{col}' must be numeric, got {df_wide[col].dtype}")


def write_machine_outputs(
    run_dirs: Dict[str, Path],
    long_df: pd.DataFrame,
    wide_df: pd.DataFrame,
    asin_scores: pd.DataFrame,
    vendor_scores: pd.DataFrame,
    signal_quality_df: Optional[pd.DataFrame] = None,
) -> Dict[str, Path]:
    machine = run_dirs["machine"]
    paths: Dict[str, Path] = {}
    # Validate schema before writing
    validate_output_schema(long_df, wide_df)
    paths["metrics_long"] = _ensure_parquet(long_df, machine / "metrics_long.parquet")
    paths["metrics_wide"] = _ensure_parquet(wide_df, machine / "metrics_wide.parquet")
    paths["scores_asin"] = _ensure_parquet(asin_scores, machine / "scores_asin.parquet")
    paths["scores_vendor"] = _ensure_parquet(vendor_scores, machine / "scores_vendor.parquet")
    if signal_quality_df is not None and not signal_quality_df.empty:
        paths["signal_quality"] = _ensure_parquet(signal_quality_df, machine / "signal_quality.parquet")
    return paths


def _auto_fit_columns(ws) -> None:
    # Simple auto width based on max length per column
    for col_cells in ws.columns:
        max_len = 0
        col = col_cells[0].column_letter
        for cell in col_cells:
            try:
                val = "" if cell.value is None else str(cell.value)
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col].width = min(max(10, max_len + 2), 60)


def write_excel_outputs(
    run_dirs: Dict[str, Path],
    vendor_scores: pd.DataFrame,
    asin_long: pd.DataFrame,
    taxonomy_df: Optional[pd.DataFrame] = None,
    anomalies_df: Optional[pd.DataFrame] = None,
) -> Path:
    from openpyxl import Workbook

    excel_dir = run_dirs["excel"]
    out_path = excel_dir / "diagnostics.xlsx"

    wb = Workbook()
    # Sheet 1: Vendor Summary
    ws1 = wb.active
    ws1.title = "Vendor Summary"
    if not vendor_scores.empty:
        ws1.append(list(vendor_scores.columns))
        for _, row in vendor_scores.iterrows():
            ws1.append(list(row.values))
    _auto_fit_columns(ws1)
    ws1.freeze_panes = "A2"

    # Sheet 2: ASIN Diagnostics (row-level)
    ws2 = wb.create_sheet("ASIN Diagnostics")
    # Validate required Patch-3 fields for Excel visibility (warnings only, do not abort)
    import warnings as _warnings
    required_excel_long = [
        "weight_effective",
        "weight_multiplier",
        "base_weight",
        "completeness",
        "recency_weeks",
        "variance",
        "trend_quality",
    ]
    missing_excel = [c for c in required_excel_long if c not in asin_long.columns]
    # Also ensure at least one metric audit set exists when wide is provided
    metric_audit_missing: List[str] = []
    metric_names = [c.split("std_")[1] for c in asin_long.columns if c.startswith("std_")]
    if metric_names:
        for m in metric_names:
            for prefix in ("contrib_", "weight_", "weightmul_", "base_weight_"):
                col = f"{prefix}{m}"
                if col not in asin_long.columns:
                    metric_audit_missing.append(col)
    if missing_excel or metric_audit_missing:
        _warnings.warn(
            f"ASIN Diagnostics input missing fields: {missing_excel + metric_audit_missing}. "
            f"Workbook will be created with available columns.")
    cols2 = list(asin_long.columns)
    if not asin_long.empty:
        ws2.append(cols2)
        for _, row in asin_long[cols2].iterrows():
            ws2.append(list(row.values))
    _auto_fit_columns(ws2)
    ws2.freeze_panes = "A2"

    # Sheet 3: Metric Definitions
    ws3 = wb.create_sheet("Metric Definitions")
    if taxonomy_df is not None and not taxonomy_df.empty:
        ws3.append(list(taxonomy_df.columns))
        for _, row in taxonomy_df.iterrows():
            ws3.append(list(row.values))
    else:
        ws3.append(["metric", "agg_rule", "is_rate", "weight_metric", "winsor_limits", "direction"])  # header only
    _auto_fit_columns(ws3)

    # Sheet 4: Flags & Anomalies
    ws4 = wb.create_sheet("Flags & Anomalies")
    if anomalies_df is not None and not anomalies_df.empty:
        ws4.append(list(anomalies_df.columns))
        for _, row in anomalies_df.iterrows():
            ws4.append(list(row.values))
    else:
        ws4.append(["note"]) ; ws4.append(["No anomalies provided"])
    _auto_fit_columns(ws4)

    wb.save(out_path)
    return out_path


# -----------------------------
# External vendor routing/mirroring
# -----------------------------

def _extract_run_tag(run_dir: Path) -> str:
    # run_dir expected like .../phase4/run_YYYYMMDD_HHMMSS
    return run_dir.name


def _ensure_dir(p: Path) -> Path:
    p.mkdir(parents=True, exist_ok=True)
    return p


def mirror_outputs_to_external(
    external_base: Path | str,
    run_dirs: Dict[str, Path],
    wide_df: pd.DataFrame,
    asin_scores: pd.DataFrame,
    vendor_scores: pd.DataFrame,
    excel_path: Optional[Path],
    run_metadata: Dict[str, object],
) -> Dict[str, List[Path]]:
    """Partition outputs per vendor and mirror to the external SEIF directory.

    Structure per vendor:
      <external_base>/<vendor_code>/<run_tag>/{asin,vendor,machine,excel,metadata}
    """
    base = Path(external_base)
    out_paths: Dict[str, List[Path]] = {}
    run_tag = _extract_run_tag(run_dirs["run"]) if run_dirs and run_dirs.get("run") else datetime.now().strftime("run_%Y%m%d_%H%M%S")

    if vendor_scores is None or vendor_scores.empty:
        # No vendors to partition; write a generic folder
        vcode = "_ALL_"
        vendor_dir = _ensure_dir(base / vcode / run_tag)
        _ensure_dir(vendor_dir / "machine")
        _ensure_dir(vendor_dir / "asin")
        _ensure_dir(vendor_dir / "vendor")
        _ensure_dir(vendor_dir / "excel")
        _ensure_dir(vendor_dir / "metadata")
        # Save generic copies
        if wide_df is not None and not wide_df.empty:
            wide_df.to_parquet(vendor_dir / "machine" / "metrics_wide.parquet", index=False)
        if asin_scores is not None and not asin_scores.empty:
            asin_scores.to_parquet(vendor_dir / "asin" / "scores_asin.parquet", index=False)
        if vendor_scores is not None and not vendor_scores.empty:
            vendor_scores.to_parquet(vendor_dir / "vendor" / "scores_vendor.parquet", index=False)
        if excel_path and excel_path.exists():
            shutil.copy2(excel_path, vendor_dir / "excel" / excel_path.name)
        # Metadata per vendor
        meta = dict(run_metadata or {})
        meta["vendor_code"] = vcode
        write_metadata_json(vendor_dir / "metadata" / "metadata.json", meta)
        out_paths[vcode] = [vendor_dir]
        return out_paths

    # Partition by vendor_code
    vendor_list = sorted(vendor_scores["vendor_code"].dropna().unique().tolist())
    for vcode in vendor_list:
        vendor_dir = _ensure_dir(base / vcode / run_tag)
        asin_dir = _ensure_dir(vendor_dir / "asin")
        vend_dir = _ensure_dir(vendor_dir / "vendor")
        mach_dir = _ensure_dir(vendor_dir / "machine")
        ex_dir = _ensure_dir(vendor_dir / "excel")
        meta_dir = _ensure_dir(vendor_dir / "metadata")

        # Filtered frames
        wv = wide_df[wide_df["vendor_code"] == vcode] if wide_df is not None and not wide_df.empty and "vendor_code" in wide_df.columns else pd.DataFrame()
        av = asin_scores[asin_scores["vendor_code"] == vcode] if asin_scores is not None and not asin_scores.empty and "vendor_code" in asin_scores.columns else pd.DataFrame()
        vv = vendor_scores[vendor_scores["vendor_code"] == vcode]

        if not wv.empty:
            wv.to_parquet(mach_dir / "metrics_wide.parquet", index=False)
        if not av.empty:
            av.to_parquet(asin_dir / "scores_asin.parquet", index=False)
        if not vv.empty:
            vv.to_parquet(vend_dir / "scores_vendor.parquet", index=False)

        if excel_path and excel_path.exists():
            # Duplicate diagnostics per vendor for convenience
            shutil.copy2(excel_path, ex_dir / excel_path.name)

        meta = dict(run_metadata or {})
        meta["vendor_code"] = vcode
        write_metadata_json(meta_dir / "metadata.json", meta)

        out_paths.setdefault(vcode, []).append(vendor_dir)

    return out_paths
