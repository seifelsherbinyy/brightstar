"""Utility helpers for Phase 5 dashboard generation."""
from __future__ import annotations

import logging
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.formatting.rule import ColorScaleRule, IconSetRule, DataBarRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .ingestion_utils import ensure_directory, load_config
from .scoring_utils import (
    build_asin_l4w_aggregates,
    add_asin_growth_segment,
    build_asin_dashboard_df,
    prepare_vendor_scoring_dashboard_df,
    prepare_asin_scoring_dashboard_df,
)


LOGGER_NAME = "brightstar.phase5"


@dataclass
class DashboardData:
    """Container holding prepared datasets used for dashboard creation."""

    normalized: pd.DataFrame
    vendor_scorecard: pd.DataFrame
    asin_scorecard: pd.DataFrame
    commentary: pd.DataFrame
    forecast_vendor: pd.DataFrame
    forecast_asin: pd.DataFrame
    forecast_health: pd.DataFrame | None = None


def setup_logging(config: Dict) -> logging.Logger:
    """Configure logging for the dashboard phase."""

    log_dir = ensure_directory(config["paths"]["logs_dir"])
    dashboard_config = config.get("dashboard", {})
    file_name = dashboard_config.get("logging_file_name", "phase5_dashboard.log")
    log_path = log_dir / file_name

    logger = logging.getLogger(LOGGER_NAME)
    logger.setLevel(getattr(logging, dashboard_config.get("logging_level", "INFO")))
    logger.handlers = []

    formatter = logging.Formatter(
        "%(asctime)s | %(levelname)s | %(name)s | %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )

    file_handler = logging.FileHandler(log_path, encoding="utf-8")
    file_handler.setFormatter(formatter)
    logger.addHandler(file_handler)

    stream_handler = logging.StreamHandler()
    stream_handler.setFormatter(formatter)
    logger.addHandler(stream_handler)

    logger.debug("Phase 5 logging initialised at %s", log_path)
    return logger


def _candidate_paths(*paths: str | Path) -> List[Path]:
    return [Path(p) for p in paths if p]


def _load_dataframe(candidates: Iterable[Path], logger: Optional[logging.Logger] = None) -> pd.DataFrame:
    """Load a dataframe from the first existing candidate path."""

    for path in candidates:
        if not path.is_file():
            continue
        suffix = path.suffix.lower()
        if suffix == ".parquet":
            try:
                return pd.read_parquet(path)
            except Exception as exc:  # pragma: no cover - optional engine
                if logger:
                    logger.warning("Unable to read parquet file %s (%s). Trying next candidate.", path, exc)
                continue
        if suffix in {".csv", ".txt"}:
            return pd.read_csv(path)
        if suffix in {".xlsx", ".xlsm"}:
            return pd.read_excel(path)
    return pd.DataFrame()


def load_processed_tables(config: Dict, logger: Optional[logging.Logger] = None) -> DashboardData:
    """Load normalized, scoring, commentary, and forecast outputs for dashboard assembly."""

    processed_dir = Path(config["paths"]["processed_dir"])

    normalized = _load_dataframe(
        _candidate_paths(
            config["paths"].get("normalized_output"),
            processed_dir / "normalized_phase1.parquet",
            processed_dir / "normalized_phase1.csv",
        ),
        logger=logger,
    )

    vendor_scorecard = _load_dataframe(
        _candidate_paths(
            config["paths"].get("scorecard_vendor_parquet"),
            config["paths"].get("scorecard_vendor_csv"),
        ),
        logger=logger,
    )

    asin_scorecard = _load_dataframe(
        _candidate_paths(
            config["paths"].get("scorecard_asin_parquet"),
            config["paths"].get("scorecard_asin_csv"),
        ),
        logger=logger,
    )

    commentary_dir = processed_dir / "commentary"
    commentary = _load_dataframe(
        _candidate_paths(
            commentary_dir / "commentary.parquet",
            commentary_dir / "commentary.csv",
        ),
        logger=logger,
    )

    forecast_dir = Path(config["paths"].get("forecasts_dir", processed_dir / "forecasts"))
    forecast_vendor = _load_dataframe(
        _candidate_paths(
            config["paths"].get("forecast_vendor_parquet"),
            forecast_dir / "vendor_forecasts.parquet",
            config["paths"].get("forecast_vendor_csv"),
            forecast_dir / "vendor_forecasts.csv",
        ),
        logger=logger,
    )
    forecast_asin = _load_dataframe(
        _candidate_paths(
            config["paths"].get("forecast_asin_parquet"),
            forecast_dir / "asin_forecasts.parquet",
            config["paths"].get("forecast_asin_csv"),
            forecast_dir / "asin_forecasts.csv",
        ),
        logger=logger,
    )

    # Optional: load evaluation health summary (Phase 6)
    eval_cfg = config.get("evaluation", {})
    eval_dir = Path(eval_cfg.get("output_dir", "outputs/forecast_evaluation"))
    health_df = _load_dataframe(
        _candidate_paths(
            eval_dir / "health_summary.parquet",
            eval_dir / "health_summary.csv",
        ),
        logger=logger,
    )

    return DashboardData(
        normalized=normalized,
        vendor_scorecard=vendor_scorecard,
        asin_scorecard=asin_scorecard,
        commentary=commentary,
        forecast_vendor=forecast_vendor,
        forecast_asin=forecast_asin,
        forecast_health=health_df if not health_df.empty else pd.DataFrame(),
    )


def _latest_commentary(commentary: pd.DataFrame, entity_type: str) -> pd.DataFrame:
    if commentary.empty:
        return pd.DataFrame(columns=["vendor_code", "asin", "metric", "Commentary_Text", "Commentary_Type"])

    filtered = commentary[commentary["entity_type"].str.lower() == entity_type.lower()].copy()
    if filtered.empty:
        return pd.DataFrame(columns=["vendor_code", "asin", "metric", "Commentary_Text", "Commentary_Type"])

    group_keys: List[str] = ["vendor_code", "metric"]
    sort_keys: List[str] = ["vendor_code", "metric", "week_label"]
    if entity_type != "vendor":
        group_keys.insert(1, "asin")
        sort_keys.insert(1, "asin")

    filtered = filtered.sort_values(by=sort_keys, ascending=[True] * (len(sort_keys) - 1) + [False])
    aggregated = filtered.groupby(group_keys, as_index=False).first()

    if entity_type == "vendor":
        aggregated["asin"] = None

    columns = ["vendor_code", "asin", "metric", "Commentary_Text", "Commentary_Type"]
    for column in columns:
        if column not in aggregated.columns:
            aggregated[column] = None
    return aggregated[columns]


def _forecast_potential(forecasts: pd.DataFrame, entity_type: str) -> pd.DataFrame:
    if forecasts.empty:
        cols = [
            "vendor_code",
            "asin",
            "metric",
            "potential_gain",
            "forecast_horizon_weeks",
            "Forecast_Risk_Flag",
            "risk_gain_flag",
            "delta_value",
            "delta_pct",
        ]
        return pd.DataFrame(columns=cols)

    filtered = forecasts[forecasts["entity_type"].str.lower() == entity_type.lower()].copy()
    if filtered.empty:
        return pd.DataFrame(
            columns=[
                "vendor_code",
                "asin",
                "metric",
                "potential_gain",
                "forecast_horizon_weeks",
                "Forecast_Risk_Flag",
                "risk_gain_flag",
                "delta_value",
                "delta_pct",
            ]
        )

    group_keys: List[str] = ["vendor_code", "metric"]
    if entity_type != "vendor":
        group_keys.insert(1, "asin")

    filtered = filtered.sort_values(
        ["forecast_horizon_weeks", "potential_gain"],
        ascending=[False, False],
    )
    aggregated = filtered.groupby(group_keys, as_index=False).first()

    if entity_type == "vendor":
        aggregated["asin"] = None
    return aggregated[
        group_keys
        + [
            "potential_gain",
            "forecast_horizon_weeks",
            "Forecast_Risk_Flag",
            "risk_gain_flag",
            "delta_value",
            "delta_pct",
        ]
    ]


def _enrich_scorecard(
    scorecard: pd.DataFrame,
    commentary: pd.DataFrame,
    forecasts: pd.DataFrame,
    entity_type: str,
) -> pd.DataFrame:
    if scorecard.empty:
        return scorecard

    enriched = scorecard.copy()
    commentary_latest = _latest_commentary(commentary, entity_type)
    forecast_summary = _forecast_potential(forecasts, entity_type)

    if entity_type == "vendor":
        commentary_latest = commentary_latest.drop(columns=["asin"], errors="ignore")
        forecast_summary = forecast_summary.drop(columns=["asin"], errors="ignore")
        enriched = enriched.merge(
            commentary_latest,
            how="left",
            on=["vendor_code", "metric"],
        )
        enriched = enriched.merge(
            forecast_summary,
            how="left",
            on=["vendor_code", "metric"],
        )
    else:
        enriched = enriched.merge(
            commentary_latest,
            how="left",
            on=["vendor_code", "asin", "metric"],
        )
        enriched = enriched.merge(
            forecast_summary,
            how="left",
            on=["vendor_code", "asin", "metric"],
        )

    if "Commentary_Type" not in enriched.columns:
        enriched["Commentary_Type"] = pd.NA
    if "Commentary_Text" not in enriched.columns:
        enriched["Commentary_Text"] = pd.NA
    enriched.rename(columns={"Commentary_Text": "Commentary_Insight"}, inplace=True)
    if "potential_gain" in enriched.columns:
        enriched.rename(columns={"potential_gain": "Forecast_Potential_Gain"}, inplace=True)
    else:
        enriched["Forecast_Potential_Gain"] = pd.NA

    rename_map = {
        "forecast_horizon_weeks": "Forecast_Horizon_Weeks",
        "risk_gain_flag": "Forecast_Risk_Label",
        "delta_value": "Forecast_Delta_Value",
        "delta_pct": "Forecast_Delta_Pct",
    }
    enriched.rename(columns=rename_map, inplace=True)

    for column in ["Forecast_Horizon_Weeks", "Forecast_Risk_Flag", "Forecast_Risk_Label", "Forecast_Delta_Value", "Forecast_Delta_Pct"]:
        if column not in enriched.columns:
            enriched[column] = pd.NA

    return enriched


def _apply_header_style(ws, formatting: Dict) -> None:
    header_fill = PatternFill("solid", fgColor=formatting.get("header_fill", "1F4E78"))
    header_font = Font(
        color=formatting.get("header_font_color", "FFFFFF"),
        bold=True,
        name=formatting.get("body_font", "Calibri"),
        size=float(formatting.get("header_font_size", 12)),
    )
    border_style = formatting.get("borders", {}).get("style", "thin")
    border_color = formatting.get("borders", {}).get("color", "D9D9D9")
    header_border = Border(
        left=Side(style=border_style, color=border_color),
        right=Side(style=border_style, color=border_color),
        top=Side(style=border_style, color=border_color),
        bottom=Side(style=border_style, color=border_color),
    ) if formatting.get("borders") else None
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        if header_border:
            cell.border = header_border


def _auto_size_columns(ws, formatting: Dict) -> None:
    default_width = float(formatting.get("column_widths", {}).get("default", 18))
    overrides = formatting.get("column_widths", {}).get("overrides", {})
    for idx, column_cells in enumerate(ws.columns, start=1):
        column_letter = get_column_letter(idx)
        header = column_cells[0].value
        target_width = overrides.get(str(header), default_width)
        ws.column_dimensions[column_letter].width = target_width


def _apply_score_conditional_formatting(ws, formatting: Dict) -> None:
    headers = [cell.value for cell in ws[1]]
    if "Composite_Score" not in headers:
        return
    column_index = headers.index("Composite_Score") + 1
    start_color = formatting.get("score_color_scale", {}).get("start_color", "F8696B")
    mid_color = formatting.get("score_color_scale", {}).get("mid_color", "FFEB84")
    end_color = formatting.get("score_color_scale", {}).get("end_color", "63BE7B")
    mid_value = float(formatting.get("score_color_scale", {}).get("mid_value", 650))
    end_value = float(formatting.get("score_color_scale", {}).get("end_value", 900))
    max_row = ws.max_row
    cell_range = f"{get_column_letter(column_index)}2:{get_column_letter(column_index)}{max_row}"
    rule = ColorScaleRule(
        start_type="num",
        start_value=0,
        start_color=start_color,
        mid_type="num",
        mid_value=mid_value,
        mid_color=mid_color,
        end_type="num",
        end_value=end_value,
        end_color=end_color,
    )
    ws.conditional_formatting.add(cell_range, rule)


def _apply_commentary_fills(ws, formatting: Dict) -> None:
    headers = [cell.value for cell in ws[1]]
    if "Commentary_Type" not in headers:
        return
    column_index = headers.index("Commentary_Type")
    color_map = formatting.get("commentary_fills", {})
    default_fill = PatternFill("solid", fgColor=color_map.get("default", "FFFFFF"))
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[column_index]
        key = str(cell.value) if cell.value is not None else ""
        fill_color = color_map.get(key, default_fill.fgColor if default_fill else "FFFFFF")
        cell.fill = PatternFill("solid", fgColor=fill_color)


def _apply_icon_sets(ws, formatting: Dict) -> None:
    icon_configs = formatting.get("icon_sets")
    if not icon_configs:
        return

    if isinstance(icon_configs, dict):
        icon_configs = [icon_configs]

    headers = [cell.value for cell in ws[1]]
    for config in icon_configs:
        column_name = config.get("column")
        if not column_name or column_name not in headers:
            continue
        col_idx = headers.index(column_name) + 1
        start_row = 2
        end_row = ws.max_row
        if end_row < start_row:
            continue
        cell_range = f"{get_column_letter(col_idx)}{start_row}:{get_column_letter(col_idx)}{end_row}"
        values = config.get("values")
        if not values:
            continue
        icon_style = config.get("style", "3Arrows")
        rule_type = config.get("type", "num")
        rule = IconSetRule(
            icon_style=icon_style,
            type=rule_type,
            values=values,
            reverse=bool(config.get("reverse", False)),
            showValue=not bool(config.get("icons_only", False)),
            percent=bool(config.get("percent", False)),
        )
        ws.conditional_formatting.add(cell_range, rule)
        logging.getLogger(LOGGER_NAME).info(
            "Applied icon set '%s' to column '%s' on sheet '%s'",
            icon_style,
            column_name,
            ws.title,
        )


def _apply_segment_colors(ws, formatting: Dict, column_name: str = "Asin_Growth_Segment") -> None:
    """Apply fill colors to the ASIN segment column using a config map."""
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        return
    col_idx = headers.index(column_name)
    seg_map = formatting.get("segment_fills", {}) or {
        "Breakout": "C6EFCE",  # green
        "Emerging": "E2F0D9",  # light green
        "Core": "D9E1F2",      # light blue
        "Squeezed": "FFF2CC",  # amber
        "Sunset": "FCE4D6",    # red tint
        "Low_Data": "EEEEEE",   # gray
        "default": "FFFFFF",
    }
    default_fill = PatternFill("solid", fgColor=seg_map.get("default", "FFFFFF"))
    for row in ws.iter_rows(min_row=2, values_only=False):
        cell = row[col_idx]
        key = str(cell.value) if cell.value is not None else ""
        color = seg_map.get(key, default_fill.fgColor if default_fill else "FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)


def _apply_databars(ws, column_name: str, color: str = "63BE7B") -> None:
    """Apply Excel data bars to a given numeric column."""
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        return
    col_idx = headers.index(column_name) + 1
    start_row = 2
    end_row = ws.max_row
    if end_row < start_row:
        return
    cell_range = f"{get_column_letter(col_idx)}{start_row}:{get_column_letter(col_idx)}{end_row}"
    rule = DataBarRule(start_type="min", end_type="max", color=color, showValue=True)
    ws.conditional_formatting.add(cell_range, rule)


def _apply_color_scale(ws, column_name: str, formatting: Dict, defaults: Dict[str, float | str] | None = None) -> None:
    """Apply a 3-color scale to a given column using dashboard score_color_scale defaults."""
    headers = [cell.value for cell in ws[1]]
    if column_name not in headers:
        return
    idx = headers.index(column_name) + 1
    max_row = ws.max_row
    if max_row < 2:
        return
    start_color = formatting.get("score_color_scale", {}).get("start_color", (defaults or {}).get("start_color", "F8696B"))
    mid_color = formatting.get("score_color_scale", {}).get("mid_color", (defaults or {}).get("mid_color", "FFEB84"))
    end_color = formatting.get("score_color_scale", {}).get("end_color", (defaults or {}).get("end_color", "63BE7B"))
    mid_value = float((defaults or {}).get("mid_value", 0))
    end_value = float((defaults or {}).get("end_value", 1))
    cell_range = f"{get_column_letter(idx)}2:{get_column_letter(idx)}{max_row}"
    rule = ColorScaleRule(
        start_type="num", start_value=0, start_color=start_color,
        mid_type="num", mid_value=mid_value, mid_color=mid_color,
        end_type="num", end_value=end_value, end_color=end_color,
    )
    ws.conditional_formatting.add(cell_range, rule)


def _apply_header_band(ws, columns: List[str], color_hex: str) -> None:
    """Apply a header fill color to a set of header columns by name."""
    headers = [cell.value for cell in ws[1]]
    fill = PatternFill("solid", fgColor=color_hex)
    for idx, name in enumerate(headers, start=1):
        if name in columns:
            ws.cell(row=1, column=idx).fill = fill


def _write_dataframe(ws, df: pd.DataFrame, formatting: Dict) -> None:
    if df.empty:
        ws.append(["No data available"])
        return

    ws.append(list(df.columns))
    precision = int(formatting.get("precision", 2))
    num_format = f"0.{''.join(['0' for _ in range(precision)])}" if precision > 0 else "0"

    body_font = Font(
        name=formatting.get("body_font", "Calibri"),
        size=float(formatting.get("body_font_size", 11)),
    )
    alignment_config = formatting.get("body_alignment", {})
    body_alignment = Alignment(
        horizontal=alignment_config.get("horizontal", "left"),
        vertical=alignment_config.get("vertical", "center"),
        wrap_text=alignment_config.get("wrap_text", True),
    )
    border_config = formatting.get("borders", {})
    border = None
    if border_config:
        border_side = Side(style=border_config.get("style", "thin"), color=border_config.get("color", "D9D9D9"))
        border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    data_row_height = float(formatting.get("row_height", 18))
    header_row_height = float(formatting.get("header_row_height", data_row_height + 2))

    for _, row in df.iterrows():
        values: List[object] = []
        for value in row.tolist():
            # Convert any pandas NA/NaN to None for openpyxl
            try:
                if pd.isna(value):
                    values.append(None)
                    continue
            except Exception:
                pass
            values.append(value)
        ws.append(values)

    _apply_header_style(ws, formatting)
    ws.row_dimensions[1].height = header_row_height
    _auto_size_columns(ws, formatting)

    headers = list(df.columns)
    for row_idx in range(2, ws.max_row + 1):
        ws.row_dimensions[row_idx].height = data_row_height
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)) and header not in {"vendor_code", "asin", "metric", "entity_type", "Commentary_Type", "Commentary_Insight"}:
                cell.number_format = num_format
            if border:
                cell.border = border
            cell.font = body_font
            cell.alignment = body_alignment

    _apply_score_conditional_formatting(ws, formatting)
    _apply_commentary_fills(ws, formatting)
    _apply_icon_sets(ws, formatting)

    # Convert used range to an Excel Table for filterable, user-friendly sheets
    try:
        if ws.max_row >= 2 and ws.max_column >= 1:
            table_name_base = "".join(ch if ch.isalnum() else "_" for ch in ws.title)
            table_name = f"{table_name_base}_Table"
            # Ensure unique name within workbook
            existing = {t.displayName for t in ws.parent._tables}
            suffix = 1
            while table_name in existing:
                table_name = f"{table_name_base}_Table_{suffix}"
                suffix += 1

            ref = f"A1:{get_column_letter(ws.max_column)}{ws.max_row}"
            table = Table(displayName=table_name, ref=ref)
            style_name = formatting.get("table_style", "TableStyleMedium2")
            table.tableStyleInfo = TableStyleInfo(
                name=style_name,
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)
            logging.getLogger(LOGGER_NAME).info(
                "Created Excel table '%s' on sheet '%s' with range %s",
                table_name,
                ws.title,
                ref,
            )
    except Exception as exc:  # pragma: no cover - defensive; table creation best-effort
        logging.getLogger(LOGGER_NAME).warning("Failed to create Excel table on sheet '%s': %s", ws.title, exc)


def _create_summary_sheet(ws, data: DashboardData, summary_config: Dict) -> Dict[str, object]:
    timestamp = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S")
    vendor_count = int(data.vendor_scorecard["vendor_code"].nunique()) if not data.vendor_scorecard.empty else 0
    if not data.vendor_scorecard.empty and "Composite_Score" in data.vendor_scorecard.columns:
        vendor_scores = (
            data.vendor_scorecard.drop_duplicates(subset=["vendor_code"])["Composite_Score"].dropna()
        )
        average_score = float(vendor_scores.mean()) if not vendor_scores.empty else 0.0
        top_score = float(vendor_scores.max()) if not vendor_scores.empty else 0.0
    else:
        average_score = 0.0
        top_score = 0.0

    if not data.forecast_vendor.empty and "potential_gain" in data.forecast_vendor.columns:
        highest_gain = float(data.forecast_vendor["potential_gain"].max())
    else:
        highest_gain = 0.0

    history_start = "Unknown"
    history_end = "Unknown"
    vendors_recent_n = None
    recent_weeks = int(summary_config.get("recent_weeks", 12))

    if not data.normalized.empty and "week_start" in data.normalized.columns:
        normalized = data.normalized.copy()
        normalized["week_start"] = pd.to_datetime(normalized["week_start"], errors="coerce")
        latest_week = normalized["week_start"].max()
        earliest_week = normalized["week_start"].min()
        freshness = latest_week.strftime("%Y-%m-%d") if pd.notna(latest_week) else "Unknown"
        history_start = earliest_week.strftime("%Y-%m-%d") if pd.notna(earliest_week) else "Unknown"
        history_end = freshness

        # Compute vendors with data in the most recent N weeks
        if pd.notna(latest_week) and "vendor_code" in normalized.columns:
            cutoff = latest_week - pd.Timedelta(weeks=recent_weeks)
            recent = normalized[normalized["week_start"] >= cutoff]
            vendors_recent_n = int(recent["vendor_code"].nunique())
    else:
        freshness = "Unknown"

    summary = {
        "Run Timestamp": timestamp,
        "Total Vendors": vendor_count,
        "Average Vendor Score": average_score,
        "Top Vendor Score": top_score,
        "Highest Forecast Gain": highest_gain,
        "Data Freshness": freshness,
        "History Start": history_start,
        "History End": history_end,
        "Recent Window (weeks)": recent_weeks,
    }

    if vendors_recent_n is not None:
        summary["Vendors Active in Recent Window"] = vendors_recent_n

    ws.title = summary_config.get("summary", "Summary")
    ws.append(["Brightstar Dashboard Summary"])
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])

    num_format = summary_config.get("summary_number_format", "#,##0.00")
    for key, value in summary.items():
        ws.append([key, value])
        cell = ws.cell(row=ws.max_row, column=2)
        if isinstance(value, float):
            cell.number_format = num_format

    for column in ("A", "B"):
        ws.column_dimensions[column].width = 28

    return summary


def create_dashboard(config: Dict, logger: logging.Logger) -> Tuple[Path, Dict[str, object]]:
    """Create the Excel dashboard workbook and return its path plus summary statistics."""

    dashboard_config = config.get("dashboard", {})
    sheet_names = dashboard_config.get("sheet_names", {})
    formatting = dashboard_config.get("formatting", {})
    output_dir = ensure_directory(dashboard_config.get("output_dir", config["paths"].get("dashboard_dir")))

    data = load_processed_tables(config, logger=logger)

    vendor_enriched = _enrich_scorecard(data.vendor_scorecard, data.commentary, data.forecast_vendor, "vendor")
    asin_enriched = _enrich_scorecard(data.asin_scorecard, data.commentary, data.forecast_asin, "asin")

    commentary_combined = data.commentary.copy()

    # Build Forecasts sheet: ensure stable schema and filter-friendly columns
    if not data.forecast_vendor.empty or not data.forecast_asin.empty:
        fv = data.forecast_vendor.copy()
        fa = data.forecast_asin.copy()

        # Standardize common columns
        def _standardize_forecast(df: pd.DataFrame, entity_type: str) -> pd.DataFrame:
            if df.empty:
                return df
            out = df.copy()
            out["entity_type"] = entity_type
            # Normalize date and value columns
            if "forecast_week_start_date" not in out.columns:
                for cand in ["week_start", "forecast_week_start", "date", "forecast_date"]:
                    if cand in out.columns:
                        out.rename(columns={cand: "forecast_week_start_date"}, inplace=True)
                        break
            if "forecast_value" not in out.columns:
                for cand in ["forecast", "value", "yhat", "prediction"]:
                    if cand in out.columns:
                        out.rename(columns={cand: "forecast_value"}, inplace=True)
                        break
            # Ensure identifier columns exist
            for col in ["vendor_code", "asin", "metric"]:
                if col not in out.columns:
                    out[col] = pd.NA
            # Ensure risk/delta/horizon columns
            rename_map = {
                "forecast_horizon_weeks": "Forecast_Horizon_Weeks",
                "risk_gain_flag": "Forecast_Risk_Label",
                "Forecast_Risk": "Forecast_Risk_Flag",
                "delta_value": "Forecast_Delta_Value",
                "delta_pct": "Forecast_Delta_Pct",
            }
            out.rename(columns=rename_map, inplace=True)
            for c in [
                "Forecast_Horizon_Weeks",
                "Forecast_Risk_Flag",
                "Forecast_Risk_Label",
                "Forecast_Delta_Value",
                "Forecast_Delta_Pct",
            ]:
                if c not in out.columns:
                    out[c] = pd.NA
            return out

        fv_std = _standardize_forecast(fv, "vendor") if not fv.empty else pd.DataFrame()
        fa_std = _standardize_forecast(fa, "asin") if not fa.empty else pd.DataFrame()

        forecasts_combined = pd.concat([fv_std, fa_std], ignore_index=True)
        # Reorder/limit columns to a stable set
        desired_cols = [
            "entity_type",
            "vendor_code",
            "asin",
            "metric",
            "forecast_week_start_date",
            "forecast_value",
            "Forecast_Horizon_Weeks",
            "Forecast_Risk_Flag",
            "Forecast_Risk_Label",
            "Forecast_Delta_Value",
            "Forecast_Delta_Pct",
        ]
        for c in desired_cols:
            if c not in forecasts_combined.columns:
                forecasts_combined[c] = pd.NA
        forecasts_combined = forecasts_combined[desired_cols]
        logger.info(
            "Forecasts sheet built with %d rows and %d columns",
            len(forecasts_combined),
            len(forecasts_combined.columns),
        )
    else:
        forecasts_combined = pd.DataFrame(
            columns=[
                "entity_type",
                "vendor_code",
                "asin",
                "metric",
                "forecast_week_start_date",
                "forecast_value",
                "Forecast_Horizon_Weeks",
                "Forecast_Risk_Flag",
                "Forecast_Risk_Label",
                "Forecast_Delta_Value",
                "Forecast_Delta_Pct",
            ]
        )

    workbook = Workbook()
    default_sheet = workbook.active
    # Pass dashboard scope config including recent_weeks and formatting to summary
    summary_stats = _create_summary_sheet(
        default_sheet,
        data,
        {**dashboard_config, **sheet_names, **formatting},
    )

    sheet_mapping = {
        sheet_names.get("vendor_scorecard", "Vendor Scorecards"): vendor_enriched,
        sheet_names.get("asin_scorecard", "ASIN Scorecards"): asin_enriched,
        sheet_names.get("commentary", "Commentary"): commentary_combined,
        sheet_names.get("forecasts", "Forecasts"): forecasts_combined,
    }

    # Build ASIN Growth sheet (precompute upstream; empty-safe)
    try:
        asin_l4w = build_asin_l4w_aggregates(data.normalized, config) if not data.normalized.empty else pd.DataFrame()
        if not asin_l4w.empty:
            asin_seg = add_asin_growth_segment(asin_l4w, config)
            asin_growth_df = build_asin_dashboard_df(asin_seg, data.forecast_asin, data.forecast_health, config)
        else:
            asin_growth_df = build_asin_dashboard_df(pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), config)
        asin_growth_sheet_name = sheet_names.get("asin_growth", "ASIN_Growth")
        sheet_mapping[asin_growth_sheet_name] = asin_growth_df
    except Exception as exc:
        logger.warning("ASIN Growth sheet construction failed; creating empty sheet. (%s)", exc)
        asin_growth_sheet_name = sheet_names.get("asin_growth", "ASIN_Growth")
        sheet_mapping[asin_growth_sheet_name] = build_asin_dashboard_df(pd.DataFrame(), pd.DataFrame(), pd.DataFrame(), config)

    # Optional Forecast Health sheet
    health_sheet_name = sheet_names.get("forecast_health", "Forecast Health")
    if hasattr(data, "forecast_health") and isinstance(data.forecast_health, pd.DataFrame) and not data.forecast_health.empty:
        sheet_mapping[health_sheet_name] = data.forecast_health.copy()

    # ------------------------------------------------------------------
    # New: Vendor & ASIN Scoring dashboard sheets (current-period KPIs)
    # ------------------------------------------------------------------
    scoring_cfg = dashboard_config.get("scoring", {})
    vendor_scoring_name = scoring_cfg.get("vendor_sheet_name", "Vendor_Scoring")
    asin_scoring_name = scoring_cfg.get("asin_sheet_name", "ASIN_Scoring")
    try:
        vendor_scoring_df = prepare_vendor_scoring_dashboard_df(config, data.vendor_scorecard, data.normalized)
    except Exception as exc:
        logger.warning("Vendor scoring DF build failed (%s); creating empty sheet.", exc)
        vendor_scoring_df = pd.DataFrame()
    try:
        asin_scoring_df = prepare_asin_scoring_dashboard_df(config, data.asin_scorecard, data.normalized)
    except Exception as exc:
        logger.warning("ASIN scoring DF build failed (%s); creating empty sheet.", exc)
        asin_scoring_df = pd.DataFrame()
    sheet_mapping[vendor_scoring_name] = vendor_scoring_df
    sheet_mapping[asin_scoring_name] = asin_scoring_df

    for title, frame in sheet_mapping.items():
        ws = workbook.create_sheet(title=title)
        _write_dataframe(ws, frame, formatting)
        logger.info("Sheet '%s' written with %d rows, %d cols", title, len(frame), len(frame.columns) if not frame.empty else 0)
        # Extra formatting for ASIN Growth sheet
        if title == asin_growth_sheet_name:
            try:
                # Freeze panes after Category (i.e., next col)
                headers = list(frame.columns) if not frame.empty else [cell.value for cell in ws[1]]
                if "Category" in headers:
                    cat_idx = headers.index("Category") + 1
                    ws.freeze_panes = f"{get_column_letter(cat_idx + 1)}2"
                # Segment colors and data bars
                _apply_segment_colors(ws, formatting, column_name="Asin_Growth_Segment")
                _apply_databars(ws, "Product_GMS_L4W", color=formatting.get("databar_color", "63BE7B"))
                # Color scales for growth delta and net margin
                _apply_color_scale(ws, "GMS_Delta_Pct", formatting, defaults={"mid_value": 0, "end_value": 0.2})
                _apply_color_scale(ws, "Net_Margin", formatting, defaults={"mid_value": 0, "end_value": 0.2})
            except Exception as exc:
                logger.warning("ASIN Growth sheet formatting partially skipped: %s", exc)
        # Formatting for Vendor/ASIN Scoring sheets
        if title in {vendor_scoring_name, asin_scoring_name}:
            try:
                # Freeze panes after identity columns
                headers = list(frame.columns) if not frame.empty else [cell.value for cell in ws[1]]
                id_cols = ["vendor_code", "vendor_name"] + (["asin", "Item_Name"] if title == asin_scoring_name else [])
                last_id_idx = max((headers.index(c) + 1 for c in id_cols if c in headers), default=2)
                ws.freeze_panes = f"{get_column_letter(last_id_idx + 1)}2"

                # Header bands per requirement (config-driven with safe defaults)
                identity_color = formatting.get("identity_header_fill", "EDEDED")  # light gray
                scores_color = formatting.get("scores_header_fill", "BDD7EE")      # blue tint
                margin_color = formatting.get("margin_header_fill", "FCE4D6")      # orange tint
                avail_color = formatting.get("availability_header_fill", "E4DFEC")  # purple tint

                score_cols = [c for c in ["Composite_Score", "Sales_Score", "Margin_Score", "Availability_Score", "Traffic_Score"] if c in headers]
                margin_cols = [c for c in ["Net_PPM"] if c in headers]
                avail_cols = [c for c in ["SoROOS", "Fill_Rate_Sourceable", "Vendor_Confirmation_Rate_Sourceable", "On_Hand_Inventory_Units_Sellable", "Total_Inventory_Units"] if c in headers]

                _apply_header_band(ws, [c for c in headers if c in id_cols], identity_color)
                _apply_header_band(ws, score_cols, scores_color)
                _apply_header_band(ws, margin_cols, margin_color)
                _apply_header_band(ws, avail_cols, avail_color)

                # Color scales
                _apply_color_scale(ws, "Composite_Score", formatting, defaults={"mid_value": 700, "end_value": 950})
                # Margin color scale
                _apply_color_scale(ws, "Net_PPM", formatting, defaults={"mid_value": 0, "end_value": 0.2})
                # Availability icon set (SoROOS traffic lights), prefer dashboard.scoring.availability_icon_set
                avail_icon = scoring_cfg.get("availability_icon_set")
                if avail_icon and "column" in avail_icon:
                    _apply_icon_sets(ws, {"icon_sets": [avail_icon]})
            except Exception as exc:
                logger.warning("Scoring sheet formatting partially skipped for '%s': %s", title, exc)

    if "Sheet" in workbook.sheetnames:
        workbook.remove(workbook["Sheet"])

    default_order = [
        sheet_names.get("summary", "Summary"),
        sheet_names.get("vendor_scorecard", "Vendor Scorecards"),
        sheet_names.get("asin_scorecard", "ASIN Scorecards"),
        vendor_scoring_name,
        asin_scoring_name,
        asin_growth_sheet_name,
        sheet_names.get("commentary", "Commentary"),
        sheet_names.get("forecasts", "Forecasts"),
    ]
    # If Forecast Health present and not explicitly ordered, append it to the end
    if health_sheet_name in sheet_mapping and health_sheet_name not in default_order:
        default_order.append(health_sheet_name)
    ordered_names = dashboard_config.get("sheet_order") or default_order
    workbook._sheets = [workbook[name] for name in ordered_names if name in workbook.sheetnames] + [
        ws for ws in workbook.worksheets if ws.title not in ordered_names
    ]

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
    file_prefix = dashboard_config.get("file_prefix", "BrightStar_Dashboard")
    output_path = Path(output_dir) / f"{file_prefix}_{timestamp}.xlsx"
    workbook.save(output_path)

    logger.info("Dashboard saved to %s", output_path)
    print(f"Dashboard created: {output_path}")

    return output_path, summary_stats


def generate_dashboard(config_path: str) -> Tuple[Path, Dict[str, object]]:
    """Orchestrate dashboard generation using the provided configuration path."""

    config = load_config(config_path)
    logger = setup_logging(config)
    return create_dashboard(config, logger)

