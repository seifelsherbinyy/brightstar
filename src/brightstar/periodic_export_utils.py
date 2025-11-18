"""Helpers for optional periodic score exports (Phase 2 parallel node).

This module implements a lightweight, offline-compatible exporter that can
slice scorecards by weekly or monthly periods and generate either:

- a single Excel workbook with one TAB per period, or
- multiple Excel files (one per period).

Neural connections (for maintainers):
    Phase 1 → (Metric Registry) → Phase 2 Scorecards → [Periodic Snapshot Output]
                                              ↘︎
                                               → Phase 5 Dashboard (independent)

The periodic export node is a parallel branch for analysis-ready snapshots.
It NEVER feeds back into scoring, commentary, forecasting, or dashboard inputs.
"""
from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Sequence, Tuple

import logging
import pandas as pd
from openpyxl import Workbook


LOGGER = logging.getLogger("brightstar.phase2")


def _safe_sheet_name(name: str) -> str:
    """Return a sheet-safe string (openpyxl constraints)."""
    # Remove/replace characters not allowed in Excel sheet names
    sanitized = "".join(ch if ch not in '[]:*?/\\' else '_' for ch in str(name))
    # Sheet name length limit is 31 chars
    return sanitized[:31] if sanitized else "Sheet"


def _week_iso_str(dt: pd.Timestamp) -> str:
    year, week, _ = dt.isocalendar()
    return f"{int(year)}-W{int(week):02d}"


def derive_period_column(df: pd.DataFrame, frequency: str) -> pd.DataFrame:
    """Attach a ``period_key`` column based on week start dates.

    - weekly: YYYY-WW (ISO week)
    - monthly: YYYY-MM

    The function attempts to locate a date-like column in typical scorecards:
      ['week_start', 'Week_Start', 'week_label', 'Week_Label'].
    If none are found or coercion fails, the function returns the input with
    an empty/NA ``period_key`` and logs a debug message (no raise).
    """
    if df is None or df.empty:
        return df.assign(period_key=pd.Series(dtype=object)) if isinstance(df, pd.DataFrame) else pd.DataFrame()

    out = df.copy()
    candidates: Sequence[str] = ("week_start", "Week_Start", "week_label", "Week_Label")
    date_col: str | None = next((c for c in candidates if c in out.columns), None)
    if date_col is None:
        # Nothing to do; degrade safely
        out["period_key"] = pd.NA
        LOGGER.info("Periodic export: no date column found on scorecard; skipping period derivation.")
        return out

    # Coerce to datetime if possible
    dates = pd.to_datetime(out[date_col], errors="coerce")
    if dates.notna().any():
        if str(frequency).lower().startswith("week"):
            out["period_key"] = dates.apply(lambda d: _week_iso_str(pd.Timestamp(d)) if pd.notna(d) else pd.NA)
        else:
            out["period_key"] = dates.dt.strftime("%Y-%m")
    else:
        out["period_key"] = pd.NA
        LOGGER.info("Periodic export: unable to parse dates from column '%s'", date_col)
    return out


def get_recent_periods(df: pd.DataFrame, max_periods: int) -> List[str]:
    """Return up to ``max_periods`` most recent distinct ``period_key`` values.

    Ordering is handled lexicographically which works for YYYY-WW and YYYY-MM.
    """
    if df is None or df.empty or "period_key" not in df.columns:
        return []
    periods = sorted(set(map(str, df["period_key"].dropna().unique().tolist())), reverse=True)
    if max_periods is None:
        return periods
    try:
        n = int(max_periods)
    except Exception:
        n = 12
    return periods[: max(0, n)]


def filter_period(df: pd.DataFrame, period_key: str) -> pd.DataFrame:
    """Return rows where ``period_key`` matches.

    Empty-safe and will return an empty DataFrame if the column is missing.
    """
    if df is None or df.empty or "period_key" not in df.columns:
        return pd.DataFrame(columns=list(df.columns) if isinstance(df, pd.DataFrame) else None)
    return df[df["period_key"] == period_key].copy()


def _write_sheet(ws, df: pd.DataFrame) -> None:
    """Write a DataFrame to a worksheet with minimal, consistent formatting.

    Reuses dashboard utilities writer if available to keep styling consistent
    across the project; otherwise falls back to a simple header+rows writer.
    """
    try:
        # Lazy import to avoid circular dep at import time
        from .dashboard_utils import _write_dataframe as dash_write

        dash_write(ws, df, formatting={})
    except Exception:
        # Minimal fallback
        if df.empty:
            ws.append(["No data available"])
            return
        ws.append(list(df.columns))
        for _, row in df.iterrows():
            ws.append([None if (isinstance(v, float) and pd.isna(v)) else v for v in row.tolist()])


def export_periods_to_tabs(scorecard_df: pd.DataFrame, periods: Iterable[str], output_path: Path) -> Path | None:
    """Create a single workbook with one sheet per period.

    Returns the created Excel file path or ``None`` if nothing was exported.
    """
    periods = [p for p in periods if p]
    if scorecard_df is None or scorecard_df.empty or not periods:
        LOGGER.info("Periodic export (tabs): nothing to export (empty data or periods).")
        return None

    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)

    wb = Workbook()
    # Replace default Sheet with first period sheet later
    for idx, period in enumerate(periods):
        ws = wb.active if idx == 0 else wb.create_sheet()
        # Prefix sheet title based on period type
        title_prefix = "Week_" if ("W" in period) else "Month_"
        ws.title = _safe_sheet_name(f"{title_prefix}{period.replace('-', '_')}")
        df_period = filter_period(scorecard_df, period)
        _write_sheet(ws, df_period)
        LOGGER.info("Periodic export (tabs): sheet '%s' rows=%d cols=%d", ws.title, len(df_period), len(df_period.columns) if not df_period.empty else 0)

    file_path = output_path if output_path.suffix.lower() == ".xlsx" else (output_path / "Periodic_Scores.xlsx")
    if file_path.is_dir():
        file_path = file_path / "Periodic_Scores.xlsx"
    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(file_path)
    LOGGER.info("Periodic export (tabs) saved: %s | periods=%d", str(file_path), len(periods))
    return file_path


def export_periods_to_files(scorecard_df: pd.DataFrame, periods: Iterable[str], output_dir: Path) -> List[Path]:
    """Create one workbook per period within ``output_dir``.

    Returns the list of generated file paths.
    """
    periods = [p for p in periods if p]
    if scorecard_df is None or scorecard_df.empty or not periods:
        LOGGER.info("Periodic export (files): nothing to export (empty data or periods).")
        return []

    output_dir = Path(output_dir)
    output_dir.mkdir(parents=True, exist_ok=True)

    paths: List[Path] = []
    for period in periods:
        wb = Workbook()
        ws = wb.active
        title_prefix = "Week_" if ("W" in period) else "Month_"
        ws.title = _safe_sheet_name(f"{title_prefix}{period.replace('-', '_')}")
        df_period = filter_period(scorecard_df, period)
        _write_sheet(ws, df_period)

        # File name hint: Vendor_Scores_<period>.xlsx
        fname = f"Vendor_Scores_{period.replace('-', '_')}.xlsx"
        fpath = output_dir / fname
        wb.save(fpath)
        LOGGER.info(
            "Periodic export (files): saved %s | rows=%d cols=%d",
            str(fpath),
            len(df_period),
            len(df_period.columns) if not df_period.empty else 0,
        )
        paths.append(fpath)

    LOGGER.info("Periodic export (files) completed: %d files", len(paths))
    return paths
